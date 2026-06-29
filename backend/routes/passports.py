"""
Passport management — list, view, download QR, delete, import/export DPP records.
"""

import csv
import io
import json
import os
import re
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, defer

from database import get_db
from models import DPPRecord
from utils import generate_qr_bytes

router = APIRouter()
CONSTRUCTASK_URL = os.getenv("CONSTRUCTASK_URL", "https://constructask.vercel.app").rstrip("/")
PUBLIC_APP_URL = os.getenv("PUBLIC_APP_URL", "http://localhost:3000").rstrip("/")

EXPORT_COLUMNS = [
    "passport_id", "product_name", "manufacturer", "category", "document_type",
    "batch_number", "origin_country", "conversion_method", "confidence_score",
    "standards_count", "properties_count", "status", "created_at",
    "description", "standards_compliance", "technical_properties",
]


def _record_to_export_row(record: DPPRecord) -> dict:
    dpp = {}
    if record.dpp_json:
        try:
            dpp = json.loads(record.dpp_json)
        except Exception:
            pass
    standards = dpp.get("standards_compliance", [])
    tech_props = dpp.get("technical_properties", {})
    tech_summary = "; ".join(
        f"{k}: {v.get('value', '')} {v.get('unit', '')}".strip()
        for k, v in tech_props.items()
    ) if isinstance(tech_props, dict) else ""

    return {
        "passport_id": record.passport_id,
        "product_name": record.product_name,
        "manufacturer": record.manufacturer,
        "category": record.category or "",
        "document_type": record.document_type or "",
        "batch_number": record.batch_number or "",
        "origin_country": record.origin_country or "",
        "conversion_method": record.conversion_method or "",
        "confidence_score": record.confidence_score or 0,
        "standards_count": record.standards_count or 0,
        "properties_count": record.properties_count or 0,
        "status": record.status or "created",
        "created_at": str(record.created_at) if record.created_at else "",
        "description": dpp.get("description", ""),
        "standards_compliance": "; ".join(standards) if isinstance(standards, list) else str(standards),
        "technical_properties": tech_summary,
    }


@router.get("/")
def list_passports(
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    total = db.query(func.count(DPPRecord.id)).scalar() or 0
    records = (
        db.query(DPPRecord)
        .options(defer(DPPRecord.qr_code_data), defer(DPPRecord.dpp_json))
        .order_by(DPPRecord.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r.id,
                "passport_id": r.passport_id,
                "product_name": r.product_name,
                "manufacturer": r.manufacturer,
                "category": r.category,
                "batch_number": r.batch_number,
                "conversion_method": r.conversion_method,
                "standards_count": r.standards_count,
                "properties_count": r.properties_count,
                "qr_code_url": f"/api/passports/{r.id}/qr",
                "constructask_qr_code_url": f"/api/passports/{r.id}/constructask-qr",
                "status": r.status,
                "created_at": str(r.created_at),
            }
            for r in records
        ],
    }


@router.get("/export/csv")
def export_csv(db: Session = Depends(get_db)):
    records = db.query(DPPRecord).order_by(DPPRecord.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for r in records:
        writer.writerow(_record_to_export_row(r))
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=dpp-passports-{date.today()}.csv"},
    )


@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    records = db.query(DPPRecord).order_by(DPPRecord.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DPP Passports"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    friendly_headers = [
        "Passport ID", "Product Name", "Manufacturer", "Category", "Document Type",
        "Batch Number", "Origin Country", "Conversion Method", "Confidence %",
        "Standards Count", "Properties Count", "Status", "Created At",
        "Description", "Standards Compliance", "Technical Properties",
    ]
    for col_idx, header in enumerate(friendly_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    data_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, record in enumerate(records, 2):
        row_data = _record_to_export_row(record)
        for col_idx, key in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = thin_border
            cell.alignment = data_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [22, 30, 25, 20, 15, 18, 15, 18, 14, 14, 14, 10, 20, 40, 45, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dpp-passports-{date.today()}.xlsx"},
    )


@router.post("/import/spreadsheet", response_model=None)
async def import_spreadsheet(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(400, "No file uploaded")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in {"csv", "xlsx", "xls"}:
        raise HTTPException(400, "Only CSV and Excel (.xlsx/.xls) files are accepted")

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 10MB)")

    try:
        if ext == "csv":
            text = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                raise HTTPException(400, "Spreadsheet has no data rows")
            raw_headers = [str(h).strip().lower().replace(" ", "_") for h in all_rows[0] if h is not None]
            rows = []
            for data_row in all_rows[1:]:
                if all(cell is None for cell in data_row):
                    continue
                row_dict = {}
                for i, header in enumerate(raw_headers):
                    row_dict[header] = data_row[i] if i < len(data_row) else None
                rows.append(row_dict)
            wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {str(e)}")

    if not rows:
        raise HTTPException(400, "No data rows found in file")

    FIELD_ALIASES = {
        "product_name": ["product_name", "product", "name", "product name"],
        "manufacturer": ["manufacturer", "mfg", "company", "manufacturer_name"],
        "category": ["category", "type", "product_category"],
        "batch_number": ["batch_number", "batch", "batch_no", "lot"],
        "origin_country": ["origin_country", "country", "origin"],
        "document_type": ["document_type", "doc_type", "doctype"],
        "description": ["description", "desc", "details"],
        "standards_compliance": ["standards_compliance", "standards", "compliance"],
        "confidence_%": ["confidence_%", "confidence_score", "confidence"],
    }

    def resolve_field(row: dict, field: str) -> str:
        aliases = FIELD_ALIASES.get(field, [field])
        for alias in aliases:
            for key in row:
                if str(key).strip().lower().replace(" ", "_") == alias:
                    val = row[key]
                    return str(val).strip() if val is not None else ""
        return ""

    imported = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows, 2):
        product_name = resolve_field(row, "product_name")
        manufacturer = resolve_field(row, "manufacturer")

        if not product_name:
            skipped += 1
            errors.append(f"Row {idx}: missing product_name — skipped")
            continue

        category = resolve_field(row, "category")
        batch_number = resolve_field(row, "batch_number")
        origin_country = resolve_field(row, "origin_country") or "India"
        doc_type = resolve_field(row, "document_type") or "tds"
        description = resolve_field(row, "description")
        standards_raw = resolve_field(row, "standards_compliance")
        standards = [s.strip() for s in standards_raw.split(";") if s.strip()] if standards_raw else []

        slug = re.sub(r"[^A-Z0-9]", "-", product_name.upper())[:20].strip("-")
        now = datetime.now(timezone.utc)
        passport_id = f"DPP-{slug}-{date.today().year}-{now.strftime('%H%M%S%f')[:8]}"

        dpp_json = {
            "dpp_version": "1.0",
            "passport_id": passport_id,
            "product_name": product_name,
            "manufacturer": manufacturer,
            "category": category,
            "description": description,
            "document_type": doc_type,
            "technical_properties": {},
            "working_properties": {},
            "application": {"primary_use": [], "suitable_for": []},
            "standards_compliance": standards,
            "packaging_and_storage": {"packaging": "", "storage": "", "shelf_life": {"value": 12, "unit": "months", "condition": ""}},
            "sustainability": {"recycled_content_pct": 0, "carbon_footprint": {"value": 0, "unit": "kgCO2e/unit"}, "recyclable": True},
            "batch_info": {"batch_number": batch_number, "production_date": str(date.today()), "origin_country": origin_country, "factory_location": ""},
            "qr_verification": {"qr_code": f"QR-{slug}-{date.today().year}", "verification_url": f"{PUBLIC_APP_URL}/?passport={passport_id}", "scan_type": "check_specification"},
            "confidence": {"overall": 90, "product_name": 95, "manufacturer": 95, "technical_properties": 85, "standards_compliance": 85},
            "source_document": {"type": "Spreadsheet Import", "document_type_code": doc_type, "document_title": product_name, "revision": "", "date_issued": "", "conversion_method": "spreadsheet_import", "converted_by": "DPP Forge", "conversion_date": str(date.today())},
            "data_rights": {"permission_status": "internal_review", "rights_holder": manufacturer, "allowed_uses": ["internal_review", "manufacturer_authorized_public_qr"], "license_notes": "Confirm manufacturer permission before external publication."},
            "evidence": {"minimum_confidence_required": 90, "field_sources": [{"field": "product_name", "source_type": "Spreadsheet Import", "source_title": file.filename, "citation": f"Row {idx}", "confidence": 95}], "quality_notes": "Spreadsheet row imported for review; add source document citations before authority-grade use."},
            "audit_trail": [{"event": "dpp_created", "actor": "DPP Forge", "method": "spreadsheet_import", "timestamp": now.isoformat()}],
        }

        qr_url = f"{PUBLIC_APP_URL}/?passport={passport_id}"
        qr_data = generate_qr_bytes(qr_url)

        record = DPPRecord(
            passport_id=passport_id,
            product_name=product_name,
            manufacturer=manufacturer,
            category=category,
            batch_number=batch_number,
            origin_country=origin_country,
            conversion_method="spreadsheet_import",
            document_type=doc_type,
            confidence_score=90.0,
            standards_count=len(standards),
            properties_count=0,
            qr_code_data=qr_data,
            dpp_json=json.dumps(dpp_json),
            status="created",
        )
        db.add(record)
        imported += 1

    if imported > 0:
        db.commit()

    return {
        "status": "ok",
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20],
        "message": f"Successfully imported {imported} passport(s)" + (f", {skipped} skipped" if skipped else ""),
    }


@router.get("/{record_id}")
def get_passport(record_id: int, db: Session = Depends(get_db)):
    record = (
        db.query(DPPRecord)
        .options(defer(DPPRecord.qr_code_data))
        .filter(DPPRecord.id == record_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Passport not found")
    return {
        "id": record.id,
        "passport_id": record.passport_id,
        "product_name": record.product_name,
        "manufacturer": record.manufacturer,
        "category": record.category,
        "batch_number": record.batch_number,
        "origin_country": record.origin_country,
        "conversion_method": record.conversion_method,
        "qr_code_url": f"/api/passports/{record.id}/qr",
        "constructask_qr_code_url": f"/api/passports/{record.id}/constructask-qr",
        "status": record.status,
        "created_at": str(record.created_at),
        "dpp_json": json.loads(record.dpp_json),
    }


@router.get("/{record_id}/qr")
def get_qr_code(record_id: int, db: Session = Depends(get_db)):
    record = db.query(DPPRecord).filter(DPPRecord.id == record_id).first()
    if not record or not record.qr_code_data:
        raise HTTPException(status_code=404, detail="QR code not found")
    return Response(
        content=record.qr_code_data,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@router.get("/{record_id}/constructask-qr")
def get_constructask_qr_code(record_id: int, db: Session = Depends(get_db)):
    record = db.query(DPPRecord).filter(DPPRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Passport not found")
    url = f"{CONSTRUCTASK_URL}/?dpp={record.passport_id}"
    return Response(
        content=generate_qr_bytes(url),
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@router.patch("/{record_id}")
def update_passport(record_id: int, payload: dict, db: Session = Depends(get_db)):
    record = db.query(DPPRecord).filter(DPPRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Passport not found")

    dpp = json.loads(record.dpp_json)

    if "product_name" in payload:
        dpp["product_name"] = payload["product_name"]
        record.product_name = payload["product_name"]
    if "manufacturer" in payload:
        dpp["manufacturer"] = payload["manufacturer"]
        record.manufacturer = payload["manufacturer"]
    if "category" in payload:
        dpp["category"] = payload["category"]
        record.category = payload["category"]
    if "description" in payload:
        dpp["description"] = payload["description"]
    if "batch_number" in payload:
        dpp.setdefault("batch_info", {})["batch_number"] = payload["batch_number"]
        record.batch_number = payload["batch_number"]
    if "origin_country" in payload:
        dpp.setdefault("batch_info", {})["origin_country"] = payload["origin_country"]
        record.origin_country = payload["origin_country"]
    if "standards_compliance" in payload:
        dpp["standards_compliance"] = payload["standards_compliance"]
        record.standards_count = len(payload["standards_compliance"])
    if "technical_properties" in payload:
        dpp["technical_properties"] = payload["technical_properties"]
        record.properties_count = len(payload["technical_properties"])
    if "dpp_json" in payload:
        dpp = payload["dpp_json"]
        record.product_name = dpp.get("product_name", record.product_name)
        record.manufacturer = dpp.get("manufacturer", record.manufacturer)
        record.category = dpp.get("category", record.category)
        record.standards_count = len(dpp.get("standards_compliance", []))
        record.properties_count = len(dpp.get("technical_properties", {}))

    record.dpp_json = json.dumps(dpp)
    db.commit()
    db.refresh(record)

    return {
        "status": "updated",
        "id": record.id,
        "passport_id": record.passport_id,
        "product_name": record.product_name,
    }


@router.get("/{record_id}/export-pdf")
def export_passport_pdf(record_id: int, db: Session = Depends(get_db)):
    record = db.query(DPPRecord).filter(DPPRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Passport not found")

    dpp = json.loads(record.dpp_json)
    pdf_bytes = _generate_dpp_pdf(dpp, record)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="DPP-{record.passport_id}.pdf"'},
    )


@router.get("/{record_id}/export-ifc")
def export_passport_ifc(record_id: int, db: Session = Depends(get_db)):
    record = db.query(DPPRecord).filter(DPPRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Passport not found")

    dpp = json.loads(record.dpp_json)
    ifc_content = _generate_ifc(dpp, record)
    return Response(
        content=ifc_content,
        media_type="application/x-step",
        headers={"Content-Disposition": f'attachment; filename="DPP-{record.passport_id}.ifc"'},
    )


def _generate_ifc(dpp: dict, record: DPPRecord) -> str:
    now = datetime.now(timezone.utc)
    ts = now.strftime("%Y-%m-%dT%H:%M:%S")
    product_name = dpp.get("product_name", "Unknown Product")
    manufacturer = dpp.get("manufacturer", "")
    category = dpp.get("category", "")
    description = dpp.get("description", "")
    passport_id = record.passport_id or ""

    lines = []
    lines.append("ISO-10303-21;")
    lines.append("HEADER;")
    lines.append(f"FILE_DESCRIPTION(('ViewDefinition [CoordinationView]'),'2;1');")
    lines.append(f"FILE_NAME('{passport_id}.ifc','{ts}',('DPP Forge'),('DPP Forge'),'','DPP Forge','');")
    lines.append("FILE_SCHEMA(('IFC4'));")
    lines.append("ENDSEC;")
    lines.append("DATA;")

    eid = 0

    def nid():
        nonlocal eid
        eid += 1
        return eid

    def ifc_str(s: str) -> str:
        return f"'{s}'" if s else "''"

    org_id = nid()
    lines.append(f"#{org_id}=IFCORGANIZATION($,{ifc_str(manufacturer)},$,$,$);")

    app_id = nid()
    lines.append(f"#{app_id}=IFCAPPLICATION(#{org_id},'1.0','DPP Forge','DPPForge');")

    person_id = nid()
    lines.append(f"#{person_id}=IFCPERSON($,$,{ifc_str(manufacturer)},$,$,$,$,$);")

    person_org_id = nid()
    lines.append(f"#{person_org_id}=IFCPERSONANDORGANIZATION(#{person_id},#{org_id},$);")

    owner_id = nid()
    lines.append(f"#{owner_id}=IFCOWNERHISTORY(#{person_org_id},#{app_id},$,.READWRITE.,$,$,$,{int(now.timestamp())});")

    ctx_id = nid()
    lines.append(f"#{ctx_id}=IFCGEOMETRICREPRESENTATIONCONTEXT($,'Model',3,1.0E-05,#0,$,$);")

    units_id = nid()
    si_len = nid()
    lines.append(f"#{si_len}=IFCSIUNIT(*,.LENGTHUNIT.,.MILLI.,.METRE.);")
    si_area = nid()
    lines.append(f"#{si_area}=IFCSIUNIT(*,.AREAUNIT.,$,.SQUARE_METRE.);")
    si_force = nid()
    lines.append(f"#{si_force}=IFCSIUNIT(*,.FORCEUNIT.,.KILO.,.NEWTON.);")
    si_pressure = nid()
    lines.append(f"#{si_pressure}=IFCSIUNIT(*,.PRESSUREUNIT.,.MEGA.,.PASCAL.);")
    lines.append(f"#{units_id}=IFCUNITASSIGNMENT((#{si_len},#{si_area},#{si_force},#{si_pressure}));")

    project_id = nid()
    lines.append(f"#{project_id}=IFCPROJECT('{_ifc_guid()}',#{owner_id},{ifc_str(passport_id)},"
                 f"{ifc_str(description)},$,$,$,(#{ctx_id}),#{units_id});")

    site_id = nid()
    lines.append(f"#{site_id}=IFCSITE('{_ifc_guid()}',#{owner_id},{ifc_str(dpp.get('batch_info', {}).get('origin_country', ''))},$,$,$,$,$,.ELEMENT.,$,$,$,$,$);")

    building_id = nid()
    lines.append(f"#{building_id}=IFCBUILDING('{_ifc_guid()}',#{owner_id},'Construction Material',$,$,$,$,$,.ELEMENT.,$,$,$);")

    rel_site = nid()
    lines.append(f"#{rel_site}=IFCRELAGGREGATES('{_ifc_guid()}',#{owner_id},$,$,#{project_id},(#{site_id}));")

    rel_bld = nid()
    lines.append(f"#{rel_bld}=IFCRELAGGREGATES('{_ifc_guid()}',#{owner_id},$,$,#{site_id},(#{building_id}));")

    mat_id = nid()
    lines.append(f"#{mat_id}=IFCMATERIAL({ifc_str(product_name)},$,{ifc_str(category)});")

    # DPP Identity property set
    props_identity = []
    for label, value in [
        ("PassportID", passport_id),
        ("ProductName", product_name),
        ("Manufacturer", manufacturer),
        ("Category", category),
        ("BatchNumber", dpp.get("batch_info", {}).get("batch_number", "")),
        ("OriginCountry", dpp.get("batch_info", {}).get("origin_country", "")),
        ("DocumentType", dpp.get("document_type", "")),
        ("ConversionMethod", record.conversion_method or ""),
        ("ConfidenceScore", str(record.confidence_score or 0)),
    ]:
        if value:
            pid = nid()
            vid = nid()
            lines.append(f"#{vid}=IFCTEXT({ifc_str(value)});")
            lines.append(f"#{pid}=IFCPROPERTYSINGLEVALUE({ifc_str(label)},$,#{vid},$);")
            props_identity.append(f"#{pid}")

    if props_identity:
        pset_id = nid()
        lines.append(f"#{pset_id}=IFCPROPERTYSET('{_ifc_guid()}',#{owner_id},'DPP_Identity',$,({','.join(props_identity)}));")
        rel_pset = nid()
        lines.append(f"#{rel_pset}=IFCRELDEFINESBYPROPERTIES('{_ifc_guid()}',#{owner_id},$,$,(#{building_id}),#{pset_id});")

    # Technical properties property set
    tech_props = dpp.get("technical_properties", {})
    if tech_props:
        props_tech = []
        for name, prop in tech_props.items():
            pid = nid()
            if isinstance(prop, dict):
                val = prop.get("value", "")
                unit = prop.get("unit", "")
                method = prop.get("test_method", "")
                display = f"{val} {unit}".strip()
                if method:
                    display += f" [{method}]"
            else:
                display = str(prop)
            vid = nid()
            lines.append(f"#{vid}=IFCTEXT({ifc_str(display)});")
            lines.append(f"#{pid}=IFCPROPERTYSINGLEVALUE({ifc_str(name)},$,#{vid},$);")
            props_tech.append(f"#{pid}")

        pset_tech_id = nid()
        lines.append(f"#{pset_tech_id}=IFCPROPERTYSET('{_ifc_guid()}',#{owner_id},'DPP_TechnicalProperties',$,({','.join(props_tech)}));")
        rel_tech = nid()
        lines.append(f"#{rel_tech}=IFCRELDEFINESBYPROPERTIES('{_ifc_guid()}',#{owner_id},$,$,(#{building_id}),#{pset_tech_id});")

    # Standards compliance property set
    standards = dpp.get("standards_compliance", [])
    if standards:
        props_std = []
        for i, std in enumerate(standards):
            pid = nid()
            vid = nid()
            lines.append(f"#{vid}=IFCTEXT({ifc_str(std)});")
            lines.append(f"#{pid}=IFCPROPERTYSINGLEVALUE('Standard_{i+1}',$,#{vid},$);")
            props_std.append(f"#{pid}")

        pset_std_id = nid()
        lines.append(f"#{pset_std_id}=IFCPROPERTYSET('{_ifc_guid()}',#{owner_id},'DPP_Standards',$,({','.join(props_std)}));")
        rel_std = nid()
        lines.append(f"#{rel_std}=IFCRELDEFINESBYPROPERTIES('{_ifc_guid()}',#{owner_id},$,$,(#{building_id}),#{pset_std_id});")

    # Sustainability property set
    sustainability = dpp.get("sustainability", {})
    if sustainability:
        props_sus = []
        recycled = sustainability.get("recycled_content_pct", 0)
        carbon = sustainability.get("carbon_footprint", {})

        if recycled:
            pid = nid()
            vid = nid()
            lines.append(f"#{vid}=IFCTEXT('{recycled}%');")
            lines.append(f"#{pid}=IFCPROPERTYSINGLEVALUE('RecycledContent',$,#{vid},$);")
            props_sus.append(f"#{pid}")

        if carbon and carbon.get("value"):
            pid = nid()
            vid = nid()
            lines.append(f"#{vid}=IFCTEXT('{carbon['value']} {carbon.get('unit', '')}');")
            lines.append(f"#{pid}=IFCPROPERTYSINGLEVALUE('CarbonFootprint',$,#{vid},$);")
            props_sus.append(f"#{pid}")

        if props_sus:
            pset_sus_id = nid()
            lines.append(f"#{pset_sus_id}=IFCPROPERTYSET('{_ifc_guid()}',#{owner_id},'DPP_Sustainability',$,({','.join(props_sus)}));")
            rel_sus = nid()
            lines.append(f"#{rel_sus}=IFCRELDEFINESBYPROPERTIES('{_ifc_guid()}',#{owner_id},$,$,(#{building_id}),#{pset_sus_id});")

    lines.append("ENDSEC;")
    lines.append("END-ISO-10303-21;")
    return "\n".join(lines)


def _ifc_guid() -> str:
    import uuid
    import base64
    raw = uuid.uuid4().bytes
    b64 = base64.b64encode(raw, altchars=b"_$").decode("ascii")
    return b64[:22]


def _generate_dpp_pdf(dpp: dict, record: DPPRecord) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm, leftMargin=20 * mm, rightMargin=20 * mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("DPPTitle", parent=styles["Heading1"], fontSize=18, textColor=HexColor("#1F2937"), spaceAfter=6)
    subtitle_style = ParagraphStyle("DPPSub", parent=styles["Normal"], fontSize=10, textColor=HexColor("#6B7280"), spaceAfter=12)
    section_style = ParagraphStyle("DPPSection", parent=styles["Heading2"], fontSize=13, textColor=HexColor("#1F2937"), spaceBefore=14, spaceAfter=6)
    body_style = ParagraphStyle("DPPBody", parent=styles["Normal"], fontSize=10, textColor=HexColor("#374151"), spaceAfter=4)
    small_style = ParagraphStyle("DPPSmall", parent=styles["Normal"], fontSize=8, textColor=HexColor("#9CA3AF"))

    elements = []
    elements.append(Paragraph("DIGITAL PRODUCT PASSPORT", small_style))
    elements.append(Paragraph(dpp.get("product_name", "Untitled Product"), title_style))
    elements.append(Paragraph(f'{dpp.get("manufacturer", "")} — {record.passport_id}', subtitle_style))

    info_data = [
        ["Category", dpp.get("category", "N/A"), "Document Type", dpp.get("document_type", "N/A")],
        ["Batch", dpp.get("batch_info", {}).get("batch_number", "N/A"), "Origin", dpp.get("batch_info", {}).get("origin_country", "N/A")],
        ["Method", record.conversion_method or "N/A", "Confidence", f"{record.confidence_score or 0:.0f}%"],
        ["Created", str(record.created_at)[:10] if record.created_at else "N/A", "Status", record.status or "active"],
    ]
    info_table = Table(info_data, colWidths=[80, 140, 80, 140])
    info_table.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (0, -1), HexColor("#6B7280")),
        ("TEXTCOLOR", (2, 0), (2, -1), HexColor("#6B7280")),
        ("TEXTCOLOR", (1, 0), (1, -1), HexColor("#1F2937")),
        ("TEXTCOLOR", (3, 0), (3, -1), HexColor("#1F2937")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)

    tech_props = dpp.get("technical_properties", {})
    if tech_props:
        elements.append(Paragraph("Technical Properties", section_style))
        tp_data = [["Property", "Value", "Unit", "Test Method"]]
        for name, prop in tech_props.items():
            if isinstance(prop, dict):
                tp_data.append([name, str(prop.get("value", "")), prop.get("unit", ""), prop.get("test_method", "")])
            else:
                tp_data.append([name, str(prop), "", ""])
        tp_table = Table(tp_data, colWidths=[140, 100, 60, 140])
        tp_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#F3F4F6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#374151")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#E5E7EB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tp_table)

    standards = dpp.get("standards_compliance", [])
    if standards:
        elements.append(Paragraph("Standards Compliance", section_style))
        for std in standards:
            elements.append(Paragraph(f"• {std}", body_style))

    desc = dpp.get("description", "")
    if desc:
        elements.append(Paragraph("Description", section_style))
        elements.append(Paragraph(desc, body_style))

    sustainability = dpp.get("sustainability", {})
    if sustainability:
        elements.append(Paragraph("Sustainability", section_style))
        carbon = sustainability.get("carbon_footprint", {})
        elements.append(Paragraph(f'Recycled Content: {sustainability.get("recycled_content_pct", 0)}%', body_style))
        if carbon:
            elements.append(Paragraph(f'Carbon Footprint: {carbon.get("value", 0)} {carbon.get("unit", "")}', body_style))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Generated by DPP Forge — {date.today()}", small_style))

    doc.build(elements)
    return buf.getvalue()


@router.delete("/{record_id}")
def delete_passport(record_id: int, db: Session = Depends(get_db)):
    if os.getenv("ENABLE_PASSPORT_DELETE", "").lower() not in {"1", "true", "yes"}:
        raise HTTPException(status_code=403, detail="Passport deletion is disabled on this deployment.")
    record = db.query(DPPRecord).filter(DPPRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Passport not found")
    db.delete(record)
    db.commit()
    return {"status": "deleted", "id": record_id, "passport_id": record.passport_id}
